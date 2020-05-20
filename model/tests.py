import numpy as np
import pathlib
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from model.dataset import Dataset
from sklearn.model_selection import KFold
from model.rp_methods import kruskal_wallis, redundancy_measure, run_pca
from model.rp_methods import performance_measurement
from model.rp_methods import bayes_classifier, support_vector_machines, \
    k_nearest_neighbors, minimum_distance_classifier, fisher_discriminant_analisys


def test_dataset(path, dataset=1, scenario=1, n_runs=3, n_subsets=3, k=3, c=1, pca=0):
    # Variables
    data = Dataset()
    runs_performance = {}

    # File Variables
    wb = load_workbook(path)
    ws = wb.active
    row = ws.max_row+1
    ws.title = 'Test Results'

    # Select test data
    data.choose_data(dataset)
    # print(data.database_selected_str, "data loaded.")

    # Pre-process data
    data.scenario_pre_processing(scenario)
    # print("Finished pre-processing data for", data.scenario_selected_str)

    # Apply Kruskal-Wallis
    data.set_feature_selection_method(2)
    data.dataset = kruskal_wallis(data.dataset, len(data.dataset['label']))
    # print("Finished applying kruskal-wallis feature selection method.")

    # Apply Correlation redundancy measure
    data.dataset, unused_label = redundancy_measure(data.dataset)
    # print("Correlation rendundancy measure applied.")
    # print("Begining tests...this might take a while")

    if pca == 1:
        data.dataset = run_pca(data.dataset, len(data.dataset['label']))

    # For all 5 classifiers
    for classifier in range(1, 6):
        # Variable to hold all runs for all classifiers
        runs_performance[classifier] = {}

        if classifier == 5:
            n_runs = int(n_runs / 5)
            n_subsets = 3

        # Run "n_runs" tests
        for run in range(0, n_runs):
            # Structure to hold results of classification
            performance = {
                'fp': 0,
                'fn': 0,
                'tp': 0,
                'tn': 0,
                'accuracy': 0,
                'avg_misclassification': 0,
                'misclassification_per_fold': [],
                'avg_misclassification_per_fold': [],
                'sensitivity': 0,
                'specificity': 0
            }

            print("run %s for classifier %s" % (str(run), str(classifier)))
            # Create dict to save run results
            runs_performance[classifier][run] = {}

            # Apply K-fold: splitting the dataset
            kf = KFold(n_splits=n_subsets, shuffle=True)

            # K-fold Executions
            for idx_train, idx_test in kf.split(data.dataset["data"], data.dataset["target"]):
                # Classification prediction
                prediction = []

                # Prepare data for training
                x_train = [data.dataset["data"][idx] for idx in idx_train]
                x_train = np.asarray(x_train).astype(np.float64)
                y_train = [data.dataset["target"][idx] for idx in idx_train]

                # Prepare data for testing
                x_test = [data.dataset["data"][idx] for idx in idx_test]
                x_test = np.asarray(x_test).astype(np.float64)
                y_test = [data.dataset["target"][idx] for idx in idx_test]

                # Minimum distance classifier (MDC)
                if classifier == 1:
                    prediction = minimum_distance_classifier(x_train, y_train, x_test, y_test)

                # Fisher Discriminant Analisys (Fisher LDA)
                elif classifier == 2:
                    prediction = fisher_discriminant_analisys(x_train, y_train, x_test, y_test)

                # K-Nearest Neighbors (KNN)
                elif classifier == 3:
                    prediction = k_nearest_neighbors(x_train, y_train, x_test, y_test, k)

                # Bayes Classifier
                elif classifier == 4:
                    prediction = bayes_classifier(x_train, y_train, x_test, y_test)

                # Support Vector Machines
                elif classifier == 5:
                    prediction = support_vector_machines(x_train, y_train, x_test, y_test, c)

                # Performance measurement
                performance = performance_measurement(y_test, prediction, scenario, performance)

            # Calculate averages
            performance['avg_misclassification'] /= n_subsets
            performance['sensitivity'] /= n_subsets
            performance['specificity'] /= n_subsets
            performance['accuracy'] /= n_subsets

            # Set Layout
            set_layout(ws, scenario)

            # Add values into the sheet
            ws.cell(column=1, row=row, value=dataset)
            ws.cell(column=2, row=row, value=run)
            ws.cell(column=3, row=row, value=classifier)
            set_values(ws, scenario, performance, row)
            row += 1

            # Save performance measurement per run
            runs_performance[classifier][run]["performance"] = performance
            runs_performance[classifier][run]["scenario"] = scenario

    # For debug
    # for classifier in runs_performance:
    #     for run in runs_performance[classifier]:
    #         print("Classifier ", classifier, " run", run)
    #         print(runs_performance[classifier][run])

    return wb


def new_sheet(scenario):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tests results'

    filename = "tests_results_scenario_%s.xlsx" % scenario
    current_dir = pathlib.Path().absolute().parent
    path = "%s/temp/%s" % (current_dir, filename)

    wb.save(path)

    return path, wb


def set_values(ws, scenario, performance, row):
    if scenario == 1:
        ws.cell(column=4, row=row, value=np.array2string(performance['fp']))
        ws.cell(column=5, row=row, value=np.array2string(performance['fn']))
        ws.cell(column=6, row=row, value=np.array2string(performance['tp']))
        ws.cell(column=7, row=row, value=np.array2string(performance['tn']))
        ws.cell(column=8, row=row, value=np.array2string(performance['accuracy']))
        ws.cell(column=9, row=row, value=np.array2string(performance['avg_misclassification']))
        ws.cell(column=10, row=row, value=np.array2string(performance['sensitivity']))
        ws.cell(column=11, row=row, value=np.array2string(performance['specificity']))

    elif scenario == 2:
        for i in range(0, 3):
            ws.cell(column=i+4, row=row, value=np.array2string(performance['fp'][i]))
            ws.cell(column=i+7, row=row, value=np.array2string(performance['fn'][i]))
            ws.cell(column=i+10, row=row, value=np.array2string(performance['tp'][i]))
            ws.cell(column=i+13, row=row, value=np.array2string(performance['tn'][i]))
            ws.cell(column=i+16, row=row, value=np.array2string(performance['accuracy'][i]))
            ws.cell(column=i+19, row=row, value=np.array2string(performance['avg_misclassification'][i]))
            ws.cell(column=i+22, row=row, value=np.array2string(performance['sensitivity'][i]))
            ws.cell(column=i+25, row=row, value=np.array2string(performance['specificity'][i]))

    else:
        for i in range(0, 18):
            ws.cell(column=i+4, row=row, value=np.array2string(performance['fp'][i]))
            ws.cell(column=i+22, row=row, value=np.array2string(performance['fn'][i]))
            ws.cell(column=i+40, row=row, value=np.array2string(performance['tp'][i]))
            ws.cell(column=i+58, row=row, value=np.array2string(performance['tn'][i]))
            ws.cell(column=i+76, row=row, value=np.array2string(performance['accuracy'][i]))
            ws.cell(column=i+94, row=row, value=np.array2string(performance['avg_misclassification'][i]))
            ws.cell(column=i+112, row=row, value=np.array2string(performance['sensitivity'][i]))
            ws.cell(column=i+130, row=row, value=np.array2string(performance['specificity'][i]))


def set_layout(ws, scenario):
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 7
    ws.column_dimensions['C'].width = 10

    ws['A1'] = 'dataset'
    ws['B1'] = 'run_id'
    ws['C1'] = 'classifier'

    if scenario == 1:
        # Sheet Layout
        for column in range(4, 8):
            letter = get_column_letter(column)
            ws.column_dimensions[letter].width = 7

        for column in range(8, 12):
            letter = get_column_letter(column)
            ws.column_dimensions[letter].width = 21

        # Column headers
        ws['D1'] = 'fp'
        ws['E1'] = 'fn'
        ws['F1'] = 'tp'
        ws['G1'] = 'tn'
        ws['H1'] = 'accuracy'
        ws['I1'] = 'avg_misclassification'
        ws['J1'] = 'sensitivity'
        ws['K1'] = 'specificity'

    elif scenario == 2:
        # Sheet Layout
        for column in range(4, 16):
            letter = get_column_letter(column)
            ws.column_dimensions[letter].width = 7

        for column in range(16, 28):
            letter = get_column_letter(column)
            ws.column_dimensions[letter].width = 21

        # Column headers
        ws['D1'] = 'fp_a'
        ws['E1'] = 'fp_b'
        ws['F1'] = 'fp_c'
        ws['G1'] = 'fn_a'
        ws['H1'] = 'fn_b'
        ws['I1'] = 'fn_c'
        ws['J1'] = 'tp_a'
        ws['K1'] = 'tp_b'
        ws['L1'] = 'tp_c'
        ws['M1'] = 'tn_a'
        ws['N1'] = 'tn_b'
        ws['O1'] = 'tn_c'
        ws['P1'] = 'accuracy_a'
        ws['Q1'] = 'accuracy_b'
        ws['R1'] = 'accuracy_c'
        ws['S1'] = 'avg_misclassification_a'
        ws['T1'] = 'avg_misclassification_b'
        ws['U1'] = 'avg_misclassification_c'
        ws['V1'] = 'sensitivity_a'
        ws['W1'] = 'sensitivity_b'
        ws['X1'] = 'sensitivity_c'
        ws['Y1'] = 'specificity_a'
        ws['Z1'] = 'specificity_b'
        ws['AA1'] = 'specificity_c'

    else:
        # Sheet Layout
        for column in range(4, 76):
            letter = get_column_letter(column)
            ws.column_dimensions[letter].width = 7

        for column in range(76, 148):
            letter = get_column_letter(column)
            ws.column_dimensions[letter].width = 21

        # Column headers
        ws['D1'] = 'fp_a'
        ws['E1'] = 'fp_b'
        ws['F1'] = 'fp_c'
        ws['G1'] = 'fp_d'
        ws['H1'] = 'fp_e'
        ws['I1'] = 'fp_f'
        ws['J1'] = 'fp_g'
        ws['K1'] = 'fp_h'
        ws['L1'] = 'fp_i'
        ws['M1'] = 'fp_j'
        ws['N1'] = 'fp_k'
        ws['O1'] = 'fp_l'
        ws['P1'] = 'fp_m'
        ws['Q1'] = 'fp_o'
        ws['R1'] = 'fp_p'
        ws['S1'] = 'fp_q'
        ws['T1'] = 'fp_r'
        ws['U1'] = 'fp_s'

        ws['V1'] = 'fn_a'
        ws['W1'] = 'fn_b'
        ws['X1'] = 'fn_c'
        ws['Y1'] = 'fn_d'
        ws['Z1'] = 'fn_e'
        ws['AA1'] = 'fn_f'
        ws['AB1'] = 'fn_g'
        ws['AC1'] = 'fn_h'
        ws['AD1'] = 'fn_i'
        ws['AE1'] = 'fn_j'
        ws['AF1'] = 'fn_k'
        ws['AG1'] = 'fn_l'
        ws['AH1'] = 'fn_m'
        ws['AI1'] = 'fn_o'
        ws['AJ1'] = 'fn_p'
        ws['AK1'] = 'fn_q'
        ws['AL1'] = 'fn_r'
        ws['AM1'] = 'fn_s'

        ws['AN1'] = 'tp_a'
        ws['AO1'] = 'tp_b'
        ws['AP1'] = 'tp_c'
        ws['AQ1'] = 'tp_d'
        ws['AR1'] = 'tp_e'
        ws['AS1'] = 'tp_f'
        ws['AT1'] = 'tp_g'
        ws['AU1'] = 'tp_h'
        ws['AV1'] = 'tp_i'
        ws['AW1'] = 'tp_j'
        ws['AX1'] = 'tp_k'
        ws['AY1'] = 'tp_l'
        ws['AZ1'] = 'tp_m'
        ws['BA1'] = 'tp_o'
        ws['BB1'] = 'tp_p'
        ws['BC1'] = 'tp_q'
        ws['BD1'] = 'tp_r'
        ws['BE1'] = 'tp_s'

        ws['BF1'] = 'tn_a'
        ws['BG1'] = 'tn_b'
        ws['BH1'] = 'tn_c'
        ws['BI1'] = 'tn_d'
        ws['BJ1'] = 'tn_e'
        ws['BK1'] = 'tn_f'
        ws['BL1'] = 'tn_g'
        ws['BM1'] = 'tn_h'
        ws['BN1'] = 'tn_i'
        ws['BO1'] = 'tn_j'
        ws['BP1'] = 'tn_k'
        ws['BQ1'] = 'tn_l'
        ws['BR1'] = 'tn_m'
        ws['BS1'] = 'tn_o'
        ws['BT1'] = 'tn_p'
        ws['BU1'] = 'tn_q'
        ws['BV1'] = 'tn_r'
        ws['BW1'] = 'tn_s'

        ws['BX1'] = 'accuracy_a'
        ws['BY1'] = 'accuracy_b'
        ws['BZ1'] = 'accuracy_c'
        ws['CA1'] = 'accuracy_d'
        ws['CB1'] = 'accuracy_e'
        ws['CC1'] = 'accuracy_f'
        ws['CD1'] = 'accuracy_g'
        ws['CE1'] = 'accuracy_h'
        ws['CF1'] = 'accuracy_i'
        ws['CG1'] = 'accuracy_j'
        ws['CH1'] = 'accuracy_k'
        ws['CI1'] = 'accuracy_l'
        ws['CJ1'] = 'accuracy_m'
        ws['CK1'] = 'accuracy_o'
        ws['CL1'] = 'accuracy_p'
        ws['CM1'] = 'accuracy_q'
        ws['CN1'] = 'accuracy_r'
        ws['CO1'] = 'accuracy_s'

        ws['CP1'] = 'avg_misclassification_a'
        ws['CQ1'] = 'avg_misclassification_b'
        ws['CR1'] = 'avg_misclassification_c'
        ws['CS1'] = 'avg_misclassification_d'
        ws['CT1'] = 'avg_misclassification_e'
        ws['CU1'] = 'avg_misclassification_f'
        ws['CV1'] = 'avg_misclassification_g'
        ws['CW1'] = 'avg_misclassification_h'
        ws['CX1'] = 'avg_misclassification_i'
        ws['CY1'] = 'avg_misclassification_j'
        ws['CZ1'] = 'avg_misclassification_k'
        ws['DA1'] = 'avg_misclassification_l'
        ws['DB1'] = 'avg_misclassification_m'
        ws['DC1'] = 'avg_misclassification_o'
        ws['DD1'] = 'avg_misclassification_p'
        ws['DE1'] = 'avg_misclassification_q'
        ws['DF1'] = 'avg_misclassification_r'
        ws['DG1'] = 'avg_misclassification_s'

        ws['DH1'] = 'sensitivity_a'
        ws['DI1'] = 'sensitivity_b'
        ws['DJ1'] = 'sensitivity_c'
        ws['DK1'] = 'sensitivity_d'
        ws['DL1'] = 'sensitivity_e'
        ws['DM1'] = 'sensitivity_f'
        ws['DN1'] = 'sensitivity_g'
        ws['DO1'] = 'sensitivity_h'
        ws['DP1'] = 'sensitivity_i'
        ws['DQ1'] = 'sensitivity_j'
        ws['DR1'] = 'sensitivity_k'
        ws['DS1'] = 'sensitivity_l'
        ws['DT1'] = 'sensitivity_m'
        ws['DU1'] = 'sensitivity_o'
        ws['DV1'] = 'sensitivity_p'
        ws['DW1'] = 'sensitivity_q'
        ws['DX1'] = 'sensitivity_r'
        ws['DY1'] = 'sensitivity_s'

        ws['DZ1'] = 'specificity_a'
        ws['EA1'] = 'specificity_b'
        ws['EB1'] = 'specificity_c'
        ws['EC1'] = 'specificity_d'
        ws['ED1'] = 'specificity_e'
        ws['EE1'] = 'specificity_f'
        ws['EF1'] = 'specificity_g'
        ws['EG1'] = 'specificity_h'
        ws['EH1'] = 'specificity_i'
        ws['EI1'] = 'specificity_j'
        ws['EJ1'] = 'specificity_k'
        ws['EK1'] = 'specificity_l'
        ws['EL1'] = 'specificity_m'
        ws['EM1'] = 'specificity_o'
        ws['EN1'] = 'specificity_p'
        ws['EO1'] = 'specificity_q'
        ws['EP1'] = 'specificity_r'
        ws['EQ1'] = 'specificity_s'
